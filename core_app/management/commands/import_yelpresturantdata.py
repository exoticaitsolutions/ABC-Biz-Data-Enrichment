import csv
from django.core.management.base import BaseCommand
from core_app.models import AbcBizYelpRestaurantData, LicenseNumber
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
class Command(BaseCommand):
    help = 'Import Yelp restaurant data from an Excel file'

    def handle(self, *args, **options):
        # Use the correct file path
        file_path = r"C:\Users\Exotica\Videos\abhishek\regauravvatsandmichaelbrewer\Abc_biz_ Yelp restaurant data - Step 3.xlsx"
        
        # Load the Excel file
        wb = load_workbook(file_path)
        sheet = wb.active  # Or specify the sheet name if necessary
        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                license_type,
                file_number,
                primary_name,
                dba_name,
                prem_addr_1,
                prem_addr_2,
                prem_city,
                prem_state,
                prem_zip,
                yelp_link,
                yelp_name,
                yelp_phone,
                yelp_web_site,
                yelp_rating,
                *extra
            ) = row[:14]

            if not yelp_name:
                continue  # Skip rows without a Yelp name

            # Normalize and validate Yelp rating
            try:
                yelp_rating = Decimal(yelp_rating) if yelp_rating else None
            except (InvalidOperation, TypeError):
                yelp_rating = None

            # Create and save the data
            AbcBizYelpRestaurantData.objects.create(
                license_type=license_type,
                file_number=file_number,
                primary_name=primary_name,
                dba_name=dba_name,
                prem_addr_1=prem_addr_1,
                prem_addr_2=prem_addr_2,
                prem_city=prem_city,
                prem_state=prem_state,
                prem_zip=prem_zip,
                yelp_link=yelp_link,
                yelp_name=yelp_name,
                yelp_phone=yelp_phone,
                yelp_web_site=yelp_web_site,
                yelp_rating=yelp_rating
            )

        # Success message
        self.stdout.write(self.style.SUCCESS('Data imported successfully'))
